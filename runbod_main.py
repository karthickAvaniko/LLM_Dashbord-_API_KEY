
# -*- coding: utf-8 -*-
"""
Avaniko AI API Gateway — Production Grade Backend
Complete implementation with all features:
- JWT auth + API key management
- Multi-user with admin/user roles
- Rate limiting + token budgets
- Usage tracking + billing
- Extraction modes (invoice/receipt/id_card)
- /v1/ask universal endpoint with intent detection
- SSE streaming responses
- ThinkFilter for <think> tags
- Multi-format file support (PDF/Image/XLSX/DOCX)
- Audit logging
"""
import os, sys, secrets, sqlite3, base64, tempfile, time, json, re, asyncio, io, logging
from datetime import datetime, timedelta
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Optional, List, Dict, Any, AsyncGenerator
from collections import defaultdict

import httpx
import bcrypt
import jwt
import uvicorn
from fastapi import FastAPI, Header, HTTPException, UploadFile, File, Form, Depends, Request, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, HTMLResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

# ════════════════════════════════════════════════════════════════════════════
# LOGGING SETUP
# ════════════════════════════════════════════════════════════════════════════

_LOG_DIR = Path("/workspace/logs")
_LOG_DIR.mkdir(parents=True, exist_ok=True)

_fmt = logging.Formatter(
    "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)

def _make_logger(name: str, log_file: str) -> logging.Logger:
    logger = logging.getLogger(name)
    logger.setLevel(logging.INFO)
    if not logger.handlers:
        ch = logging.StreamHandler(sys.stdout)
        ch.setFormatter(_fmt)
        fh = RotatingFileHandler(
            _LOG_DIR / log_file, maxBytes=10 * 1024 * 1024, backupCount=5, encoding="utf-8"
        )
        fh.setFormatter(_fmt)
        logger.addHandler(ch)
        logger.addHandler(fh)
    return logger

log      = _make_logger("avaniko",       "backend.log")    # general app log
req_log  = _make_logger("avaniko.req",   "requests.log")   # per-request log
ai_log   = _make_logger("avaniko.ai",    "ai.log")         # AI calls log
err_log  = _make_logger("avaniko.error", "errors.log")     # errors only

# ════════════════════════════════════════════════════════════════════════════
# CONFIG
# ════════════════════════════════════════════════════════════════════════════

_ENV = Path("/workspace/.env")
if _ENV.exists():
    for line in _ENV.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


class Settings:
    APP_ENV = os.getenv("APP_ENV", "development").lower()
    PROJECT_NAME = "Avaniko AI API Gateway"
    VERSION = "2.0.0"

    VLLM_URL = os.getenv("VLLM_URL", "http://localhost:8000")
    MODEL_NAME = os.getenv("MODEL_NAME", "qwen3.6-35b")
    MODEL_DISPLAY_NAME = os.getenv("MODEL_DISPLAY_NAME", "Qwen3.6-35B-A3B")

    WHISPER_MODEL = os.getenv("WHISPER_MODEL", "large-v3")
    WHISPER_DEVICE = os.getenv("WHISPER_DEVICE", "cuda")

    JWT_SECRET = os.getenv("JWT_SECRET") or secrets.token_urlsafe(48)
    JWT_ALGORITHM = "HS256"
    JWT_EXPIRY_DAYS = int(os.getenv("JWT_EXPIRY_DAYS", "7"))
    ADMIN_SECRET = os.getenv("ADMIN_SECRET") or secrets.token_urlsafe(32)

    ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "karthick.murugan@avaniko.com")
    ADMIN_NAME = os.getenv("ADMIN_NAME", "Karthick Murugan")
    ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "Avan@123")

    DB_PATH = os.getenv("DB_PATH", "/workspace/backend/storage/gateway.db")
    DIST_DIR = os.getenv("DIST_DIR", "/workspace/dist")

    PRICE_PROMPT_1K = float(os.getenv("PRICE_PROMPT_1K", "0.0015"))
    PRICE_COMPLETION_1K = float(os.getenv("PRICE_COMPLETION_1K", "0.0020"))

    DEFAULT_RATE_PER_MIN = int(os.getenv("DEFAULT_RATE_PER_MIN", "10"))
    DEFAULT_RATE_PER_DAY = int(os.getenv("DEFAULT_RATE_PER_DAY", "500"))
    DEFAULT_TOKEN_BUDGET = int(os.getenv("DEFAULT_TOKEN_BUDGET", "1000000"))

    MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "25"))
    MAX_MESSAGES = int(os.getenv("MAX_MESSAGES_PER_REQUEST", "50"))
    REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "180"))
    LONG_DOC_THRESHOLD = int(os.getenv("LONG_DOC_THRESHOLD", "80000"))


settings = Settings()
print(f"[Avaniko] {settings.PROJECT_NAME} v{settings.VERSION}")
print(f"[Avaniko] APP_ENV={settings.APP_ENV} | Model={settings.MODEL_DISPLAY_NAME}")
print(f"[Avaniko] vLLM={settings.VLLM_URL}")

# ════════════════════════════════════════════════════════════════════════════
# DATABASE
# ════════════════════════════════════════════════════════════════════════════

def get_db():
    os.makedirs(os.path.dirname(settings.DB_PATH), exist_ok=True)
    conn = sqlite3.connect(settings.DB_PATH, check_same_thread=False, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn


DEFAULT_MODES = [
    {
        "name": "invoice",
        "display_name": "GST Tax Invoice",
        "description": "Extract structured data from Indian GST tax invoices",
        "system_prompt": """You are an expert at extracting data from Indian GST Tax Invoices.
Extract ALL data from the invoice image into the exact JSON schema below.
Use null for missing fields. Numbers must be numeric (not strings). Dates in DD/MM/YYYY format.
Return ONLY valid JSON, no markdown, no explanations.""",
        "json_schema": json.dumps({
            "invoice_type": "string",
            "irn": "string",
            "ack_no": "string",
            "ack_date": "string",
            "vendor": {"name": "string", "gstin": "string", "address": "string", "state": "string", "state_code": "string"},
            "bill_to": {"name": "string", "gstin": "string", "address": "string", "state": "string", "state_code": "string"},
            "ship_to": {"name": "string", "gstin": "string", "address": "string", "state": "string", "state_code": "string"},
            "invoice_number": "string",
            "invoice_date": "string",
            "po_number": "string",
            "items": [{"description": "string", "hsn_sac": "string", "qty": "number", "unit": "string", "rate": "number", "amount": "number", "cgst_rate": "number", "cgst_amount": "number", "sgst_rate": "number", "sgst_amount": "number", "igst_rate": "number", "igst_amount": "number"}],
            "subtotal": "number", "total_cgst": "number", "total_sgst": "number", "total_igst": "number",
            "round_off": "number", "grand_total": "number", "amount_in_words": "string"
        }),
        "temperature": 0.0,
        "max_tokens": 4096
    },
    {
        "name": "receipt",
        "display_name": "Retail Receipt",
        "description": "Extract data from retail shop receipts and restaurant bills",
        "system_prompt": """Extract data from this retail receipt into JSON.
Return ONLY valid JSON. Use null for missing fields.""",
        "json_schema": json.dumps({
            "merchant_name": "string", "merchant_address": "string", "date": "string", "time": "string",
            "items": [{"name": "string", "qty": "number", "price": "number"}],
            "subtotal": "number", "tax": "number", "total": "number", "payment_method": "string"
        }),
        "temperature": 0.0,
        "max_tokens": 2048
    },
    {
        "name": "id_card",
        "display_name": "ID Card",
        "description": "Extract data from Indian ID cards (Aadhaar, PAN, Passport, DL)",
        "system_prompt": """Extract data from this Indian ID card (Aadhaar/PAN/Passport/Driving License).
Identify card type first. Return ONLY valid JSON.""",
        "json_schema": json.dumps({
            "card_type": "string",
            "name": "string", "id_number": "string", "dob": "string", "gender": "string",
            "father_name": "string", "address": "string", "issue_date": "string", "expiry_date": "string"
        }),
        "temperature": 0.0,
        "max_tokens": 1024
    },
    {
        "name": "code",
        "display_name": "Code Assistant",
        "description": "Write, review, debug or explain code in any language",
        "system_prompt": """You are an expert software engineer. Write clean, efficient, well-commented code.
When writing code: provide complete working implementations with brief explanations.
When reviewing/debugging: identify issues clearly and provide corrected code.
Use markdown code blocks with the correct language tag. Be concise but thorough.""",
        "json_schema": None,
        "temperature": 0.2,
        "max_tokens": 8192
    },
    {
        "name": "think",
        "display_name": "Deep Reasoning",
        "description": "Extended thinking for complex reasoning, math and logic tasks",
        "system_prompt": """You are a careful analytical reasoner. Think step-by-step through complex problems.
Show your reasoning process clearly. For math/logic problems: show all steps.
For ambiguous problems: state your assumptions explicitly.
Provide a clear, well-structured final answer.""",
        "json_schema": None,
        "temperature": 0.6,
        "max_tokens": 16384
    },
    {
        "name": "summarize",
        "display_name": "Summarizer",
        "description": "Concise summaries of documents, articles or long text",
        "system_prompt": """You are an expert summarizer. Produce clear, concise summaries that capture all key points.
Preserve important facts, numbers, and names. Use bullet points for readability.
Output should be significantly shorter than the original. Do NOT add external information.""",
        "json_schema": None,
        "temperature": 0.3,
        "max_tokens": 2048
    },
    {
        "name": "translate",
        "display_name": "Translator",
        "description": "Accurate translation between any languages",
        "system_prompt": """You are a professional translator. Translate the given text accurately while preserving the original tone, style and meaning.
Keep proper nouns, brand names, and technical terms appropriate. Maintain formatting.
Output ONLY the translated text — no explanations unless specifically asked.""",
        "json_schema": None,
        "temperature": 0.1,
        "max_tokens": 4096
    },
    {
        "name": "json_extract",
        "display_name": "JSON Extractor",
        "description": "Extract any structured data from images or text as clean JSON",
        "system_prompt": """You are a precise data extraction engine. Extract ALL structured information from the input.
Output ONLY valid JSON — no markdown fences, no explanations, no extra text.
Use null for missing fields. Preserve numbers as numbers (not strings).
Create a logical JSON structure that best represents the content.""",
        "json_schema": None,
        "temperature": 0.0,
        "max_tokens": 4096
    }
]


def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT DEFAULT 'user',
            is_active INTEGER DEFAULT 1,
            created_at TEXT NOT NULL,
            last_login TEXT
        );

        CREATE TABLE IF NOT EXISTS api_keys (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            key TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            created_at TEXT NOT NULL,
            last_used_at TEXT,
            expires_at TEXT,
            is_active INTEGER DEFAULT 1,
            rate_limit_per_min INTEGER DEFAULT 10,
            rate_limit_per_day INTEGER DEFAULT 500,
            token_budget INTEGER DEFAULT 1000000,
            tokens_used INTEGER DEFAULT 0,
            total_requests INTEGER DEFAULT 0,
            total_cost REAL DEFAULT 0,
            allowed_endpoints TEXT DEFAULT '*',
            FOREIGN KEY (user_id) REFERENCES users(id)
        );

        CREATE INDEX IF NOT EXISTS idx_keys_key ON api_keys(key);
        CREATE INDEX IF NOT EXISTS idx_keys_user ON api_keys(user_id);

        CREATE TABLE IF NOT EXISTS usage_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            api_key TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            endpoint TEXT NOT NULL,
            mode TEXT,
            prompt_tokens INTEGER DEFAULT 0,
            completion_tokens INTEGER DEFAULT 0,
            total_tokens INTEGER DEFAULT 0,
            cost REAL DEFAULT 0,
            latency_ms INTEGER DEFAULT 0,
            status TEXT DEFAULT 'success',
            error_message TEXT,
            ip_address TEXT,
            timestamp TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_logs_key ON usage_logs(api_key);
        CREATE INDEX IF NOT EXISTS idx_logs_user ON usage_logs(user_id);
        CREATE INDEX IF NOT EXISTS idx_logs_ts ON usage_logs(timestamp);

        CREATE TABLE IF NOT EXISTS extraction_modes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            display_name TEXT NOT NULL,
            description TEXT,
            system_prompt TEXT NOT NULL,
            json_schema TEXT,
            temperature REAL DEFAULT 0.0,
            max_tokens INTEGER DEFAULT 2048,
            is_active INTEGER DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT
        );

        CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            action TEXT NOT NULL,
            details TEXT,
            ip_address TEXT,
            timestamp TEXT NOT NULL
        );
    """)

    # Seed admin
    admin = conn.execute("SELECT id FROM users WHERE email=?", (settings.ADMIN_EMAIL,)).fetchone()
    if not admin:
        pw = bcrypt.hashpw(settings.ADMIN_PASSWORD.encode(), bcrypt.gensalt()).decode()
        conn.execute(
            "INSERT INTO users (email, name, password_hash, role, created_at) VALUES (?,?,?,?,?)",
            (settings.ADMIN_EMAIL, settings.ADMIN_NAME, pw, 'admin', datetime.utcnow().isoformat())
        )
        print(f"[DB] Admin created: {settings.ADMIN_EMAIL}")

    # Seed default extraction modes
    for mode in DEFAULT_MODES:
        existing = conn.execute("SELECT id FROM extraction_modes WHERE name=?", (mode["name"],)).fetchone()
        if not existing:
            conn.execute("""
                INSERT INTO extraction_modes (name, display_name, description, system_prompt, json_schema, temperature, max_tokens, created_at)
                VALUES (?,?,?,?,?,?,?,?)
            """, (mode["name"], mode["display_name"], mode["description"], mode["system_prompt"],
                  mode["json_schema"], mode["temperature"], mode["max_tokens"], datetime.utcnow().isoformat()))
            print(f"[DB] Seeded mode: {mode['name']}")

    conn.commit()
    conn.close()


# ════════════════════════════════════════════════════════════════════════════
# SECURITY
# ════════════════════════════════════════════════════════════════════════════

def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_jwt(user_id: int, email: str, role: str) -> str:
    return jwt.encode({
        "user_id": user_id, "email": email, "role": role,
        "iat": datetime.utcnow(),
        "exp": datetime.utcnow() + timedelta(days=settings.JWT_EXPIRY_DAYS)
    }, settings.JWT_SECRET, algorithm=settings.JWT_ALGORITHM)


def decode_jwt(token: str):
    try:
        return jwt.decode(token, settings.JWT_SECRET, algorithms=[settings.JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")


def generate_api_key() -> str:
    return "ak_" + secrets.token_urlsafe(32)


# ════════════════════════════════════════════════════════════════════════════
# RATE LIMITING
# ════════════════════════════════════════════════════════════════════════════

_rate_min = defaultdict(list)
_rate_day = defaultdict(list)


def check_rate_limit(key: str, per_min: int, per_day: int):
    now = time.time()
    _rate_min[key] = [t for t in _rate_min[key] if t > now - 60]
    _rate_day[key] = [t for t in _rate_day[key] if t > now - 86400]
    if len(_rate_min[key]) >= per_min:
        raise HTTPException(429, f"Rate limit: {per_min}/min exceeded")
    if len(_rate_day[key]) >= per_day:
        raise HTTPException(429, f"Daily limit: {per_day}/day exceeded")
    _rate_min[key].append(now)
    _rate_day[key].append(now)


# ════════════════════════════════════════════════════════════════════════════
# AUTH DEPENDENCIES
# ════════════════════════════════════════════════════════════════════════════

def get_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Missing Bearer token")
    return decode_jwt(authorization[7:].strip())


def get_admin(user: dict = Depends(get_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    return user


def get_key_info(x_api_key: Optional[str] = Header(None)) -> dict:
    if not x_api_key or not x_api_key.startswith("ak_"):
        raise HTTPException(401, "Missing or invalid x-api-key header")
    conn = get_db()
    row = conn.execute("""
        SELECT ak.*, u.email as user_email, u.name as user_name, u.role as user_role
        FROM api_keys ak JOIN users u ON ak.user_id = u.id
        WHERE ak.key = ? AND ak.is_active = 1 AND u.is_active = 1
    """, (x_api_key,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(401, "Invalid or inactive API key")
    info = dict(row)

    # Expiry check
    if info.get("expires_at"):
        try:
            if datetime.fromisoformat(info["expires_at"]) < datetime.utcnow():
                raise HTTPException(401, "API key expired")
        except (ValueError, TypeError):
            pass

    # Token budget check
    if info["tokens_used"] >= info["token_budget"]:
        raise HTTPException(429, "Token budget exhausted")

    check_rate_limit(x_api_key, info["rate_limit_per_min"], info["rate_limit_per_day"])
    return info


# ════════════════════════════════════════════════════════════════════════════
# LOGGING
# ════════════════════════════════════════════════════════════════════════════

def log_usage(key: str, user_id: int, endpoint: str, pt: int, ct: int,
              latency_ms: int = 0, status: str = "success", error: str = None,
              ip: str = None, mode: str = None):
    cost = (pt / 1000 * settings.PRICE_PROMPT_1K) + (ct / 1000 * settings.PRICE_COMPLETION_1K)
    conn = get_db()
    conn.execute("""
        INSERT INTO usage_logs (api_key, user_id, endpoint, mode, prompt_tokens, completion_tokens,
                                total_tokens, cost, latency_ms, status, error_message, ip_address, timestamp)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (key, user_id, endpoint, mode, pt, ct, pt + ct, cost, latency_ms, status, error, ip,
          datetime.utcnow().isoformat()))
    conn.execute("""
        UPDATE api_keys SET last_used_at=?, total_requests=total_requests+1,
                            total_cost=total_cost+?, tokens_used=tokens_used+?
        WHERE key=?
    """, (datetime.utcnow().isoformat(), cost, pt + ct, key))
    conn.commit()
    conn.close()


def log_audit(user_id: Optional[int], action: str, details: str = None, ip: str = None):
    conn = get_db()
    conn.execute(
        "INSERT INTO audit_logs (user_id, action, details, ip_address, timestamp) VALUES (?,?,?,?,?)",
        (user_id, action, details, ip, datetime.utcnow().isoformat())
    )
    conn.commit()
    conn.close()


# ════════════════════════════════════════════════════════════════════════════
# THINK FILTER (strip <think>...</think> from streaming response)
# ════════════════════════════════════════════════════════════════════════════

class ThinkFilter:
    """Strip <think>...</think> tags from streaming text, even across chunks."""
    def __init__(self):
        self.buf = ""
        self.in_think = False

    def feed(self, chunk: str) -> str:
        self.buf += chunk
        out = []
        while self.buf:
            if self.in_think:
                idx = self.buf.find("</think>")
                if idx == -1:
                    return "".join(out)
                self.buf = self.buf[idx + len("</think>"):]
                self.in_think = False
            else:
                idx = self.buf.find("<think>")
                if idx == -1:
                    # Hold back last few chars in case "<think>" is split across chunks
                    if len(self.buf) > 8:
                        out.append(self.buf[:-8])
                        self.buf = self.buf[-8:]
                    return "".join(out)
                out.append(self.buf[:idx])
                self.buf = self.buf[idx + len("<think>"):]
                self.in_think = True
        return "".join(out)

    def flush(self) -> str:
        return "" if self.in_think else self.buf


def strip_think(text: str) -> str:
    return re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()


# ════════════════════════════════════════════════════════════════════════════
# WHISPER (lazy)
# ════════════════════════════════════════════════════════════════════════════

_whisper = None
def get_whisper():
    global _whisper
    if _whisper is None:
        from faster_whisper import WhisperModel
        print(f"[Whisper] Loading {settings.WHISPER_MODEL}...")
        _whisper = WhisperModel(settings.WHISPER_MODEL, device=settings.WHISPER_DEVICE,
                                compute_type="float16", download_root="/workspace/whisper-models")
        print("[Whisper] Ready")
    return _whisper


# ════════════════════════════════════════════════════════════════════════════
# FILE PARSERS
# ════════════════════════════════════════════════════════════════════════════

async def file_to_images_b64(data: bytes, mime: str, filename: str = "") -> List[Dict[str, str]]:
    """Convert uploaded file to list of base64 image data URLs for vLLM vision."""
    ext = (filename.split(".")[-1] if "." in filename else "").lower()

    # Direct image
    if mime.startswith("image/") or ext in ("png", "jpg", "jpeg", "webp", "gif", "bmp"):
        b64 = base64.b64encode(data).decode()
        return [{"url": f"data:{mime or 'image/jpeg'};base64,{b64}"}]

    # PDF — convert each page to image
    if mime == "application/pdf" or ext == "pdf":
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(stream=data, filetype="pdf")
            images = []
            for page in doc:
                pix = page.get_pixmap(dpi=200)
                img_bytes = pix.tobytes("png")
                b64 = base64.b64encode(img_bytes).decode()
                images.append({"url": f"data:image/png;base64,{b64}"})
            doc.close()
            return images
        except ImportError:
            raise HTTPException(500, "PyMuPDF not installed. Run: pip install pymupdf")

    raise HTTPException(400, f"Unsupported file type: {mime} ({ext})")


def file_to_text(data: bytes, filename: str) -> str:
    """Extract text from non-image files (xlsx, docx, csv, txt)."""
    ext = (filename.split(".")[-1] if "." in filename else "").lower()
    if ext == "txt" or ext == "csv":
        return data.decode("utf-8", errors="ignore")
    if ext == "xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True)
            out = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                out.append(f"=== Sheet: {sheet} ===")
                for row in ws.iter_rows(values_only=True):
                    out.append("\t".join(str(c) if c is not None else "" for c in row))
            return "\n".join(out)
        except ImportError:
            raise HTTPException(500, "openpyxl not installed")
    if ext == "docx":
        try:
            from docx import Document
            doc = Document(io.BytesIO(data))
            return "\n".join(p.text for p in doc.paragraphs)
        except ImportError:
            raise HTTPException(500, "python-docx not installed")
    raise HTTPException(400, f"Cannot extract text from .{ext}")


# ════════════════════════════════════════════════════════════════════════════
# INTENT DETECTION
# ════════════════════════════════════════════════════════════════════════════

INTENT_KEYWORDS = {
    "invoice": ["invoice", "gst", "bill", "tax invoice", "irn", "hsn", "gstin"],
    "receipt": ["receipt", "purchase", "shop bill", "restaurant", "merchant"],
    "id_card": ["aadhaar", "aadhar", "pan card", "passport", "driving license", "id card", "identity"],
}


def detect_intent(prompt: str, has_file: bool) -> Optional[str]:
    if not has_file:
        return None
    p = prompt.lower()
    for mode, keywords in INTENT_KEYWORDS.items():
        if any(kw in p for kw in keywords):
            return mode
    return None


# ════════════════════════════════════════════════════════════════════════════
# VLLM CLIENT
# ════════════════════════════════════════════════════════════════════════════

async def vllm_chat(messages: list, max_tokens: int = 2048, temperature: float = 0.0,
                    enable_thinking: bool = False) -> dict:
    payload = {
        "model": settings.MODEL_NAME,
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": temperature,
    }
    if enable_thinking:
        payload["chat_template_kwargs"] = {"enable_thinking": True}
    t0 = time.time()
    async with httpx.AsyncClient(timeout=settings.REQUEST_TIMEOUT) as client:
        resp = await client.post(f"{settings.VLLM_URL}/v1/chat/completions", json=payload)
        resp.raise_for_status()
        result = resp.json()
    u = result.get("usage", {})
    ai_log.info("chat  pt=%d ct=%d temp=%.2f think=%s latency=%dms",
                u.get("prompt_tokens", 0), u.get("completion_tokens", 0),
                temperature, enable_thinking, int((time.time() - t0) * 1000))
    return result


async def vllm_stream(messages: list, max_tokens: int = 2048, temperature: float = 0.0,
                      enable_thinking: bool = False) -> AsyncGenerator[str, None]:
    """Legacy: yields raw content strings. Use vllm_stream_events for full SSE."""
    async for event_type, text, _ in vllm_stream_events(messages, max_tokens, temperature, enable_thinking):
        if event_type == "delta":
            yield text


async def vllm_stream_events(messages: list, max_tokens: int = 2048, temperature: float = 0.0,
                              enable_thinking: bool = False):
    """
    Yields (event_type, text, usage) tuples.
    event_type: 'delta' | 'thinking' | 'done' | 'error'
    - delta:   (text_chunk, None)
    - thinking:(thinking_chunk, None)
    - done:    (finish_reason, usage_dict)
    - error:   (error_str, None)
    """
    payload = {
        "model": settings.MODEL_NAME,
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": temperature,
        "stream": True,
        "stream_options": {"include_usage": True},
    }
    if enable_thinking:
        payload["chat_template_kwargs"] = {"enable_thinking": True}

    tf = ThinkFilter()
    final_usage = {}
    finish_reason = None

    try:
        async with httpx.AsyncClient(timeout=settings.REQUEST_TIMEOUT) as client:
            async with client.stream("POST", f"{settings.VLLM_URL}/v1/chat/completions", json=payload) as resp:
                resp.raise_for_status()
                async for line in resp.aiter_lines():
                    if not line.startswith("data: "):
                        continue
                    data = line[6:]
                    if data == "[DONE]":
                        break
                    try:
                        obj = json.loads(data)
                        if obj.get("usage"):
                            final_usage = obj["usage"]
                        choices = obj.get("choices") or []
                        if not choices:
                            continue
                        c = choices[0]
                        if c.get("finish_reason"):
                            finish_reason = c["finish_reason"]
                        delta = c.get("delta") or {}
                        # Qwen3 thinking via reasoning_content field
                        rc = delta.get("reasoning_content") or ""
                        if rc and enable_thinking:
                            yield ("thinking", rc, None)
                        # Main content — strip inline <think> tags
                        content = delta.get("content") or ""
                        if content:
                            clean = tf.feed(content)
                            if clean:
                                yield ("delta", clean, None)
                    except json.JSONDecodeError:
                        continue

        tail = tf.flush()
        if tail:
            yield ("delta", tail, None)
        yield ("done", finish_reason or "stop", final_usage)

    except Exception as e:
        yield ("error", str(e)[:300], None)


# ════════════════════════════════════════════════════════════════════════════
# FASTAPI APP
# ════════════════════════════════════════════════════════════════════════════

app = FastAPI(title=settings.PROJECT_NAME, version=settings.VERSION)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_origin_regex=".*",
    allow_methods=["*"], allow_headers=["*"], expose_headers=["*"],
    allow_credentials=False, max_age=600,
)


@app.middleware("http")
async def request_logger(request: Request, call_next):
    start = time.time()
    try:
        response = await call_next(request)
        latency = int((time.time() - start) * 1000)
        req_log.info(
            "%s %s %s %dms  ip=%s",
            request.method, request.url.path,
            response.status_code, latency,
            request.client.host if request.client else "-"
        )
        return response
    except Exception as exc:
        latency = int((time.time() - start) * 1000)
        err_log.exception("Unhandled %s %s %dms: %s", request.method, request.url.path, latency, exc)
        raise


@app.options("/{full_path:path}", include_in_schema=False)
async def preflight(full_path: str, request: Request):
    origin = request.headers.get("origin", "*")
    return Response(status_code=204, headers={
        "Access-Control-Allow-Origin": origin,
        "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS, PATCH",
        "Access-Control-Allow-Headers": "*",
    })


@app.on_event("startup")
def startup():
    init_db()
    log.info("=" * 60)
    log.info("%s v%s started", settings.PROJECT_NAME, settings.VERSION)
    log.info("Model: %s", settings.MODEL_DISPLAY_NAME)
    log.info("vLLM:  %s", settings.VLLM_URL)
    log.info("DB:    %s", settings.DB_PATH)
    log.info("Logs:  %s", str(_LOG_DIR))
    log.info("=" * 60)


# ════════════════════════════════════════════════════════════════════════════
# PUBLIC
# ════════════════════════════════════════════════════════════════════════════

@app.get("/")
def root():
    return {
        "service": settings.PROJECT_NAME, "version": settings.VERSION,
        "model": settings.MODEL_DISPLAY_NAME, "status": "running",
        "docs": "/docs"
    }


@app.get("/health")
async def health():
    vllm_ok = False
    try:
        async with httpx.AsyncClient(timeout=5) as client:
            r = await client.get(f"{settings.VLLM_URL}/v1/models")
            vllm_ok = r.status_code == 200
    except Exception:
        pass
    return {
        "status": "ok" if vllm_ok else "degraded",
        "vllm": "online" if vllm_ok else "offline",
        "model": settings.MODEL_DISPLAY_NAME,
        "timestamp": datetime.utcnow().isoformat()
    }


# ════════════════════════════════════════════════════════════════════════════
# AUTH ROUTES
# ════════════════════════════════════════════════════════════════════════════

class RegisterReq(BaseModel):
    email: str = Field(..., min_length=5, max_length=200)
    name: str = Field(..., min_length=2, max_length=100)
    password: str = Field(..., min_length=8, max_length=128)


class LoginReq(BaseModel):
    email: str
    password: str


@app.post("/auth/register", status_code=201)
def register(req: RegisterReq, request: Request):
    conn = get_db()
    existing = conn.execute("SELECT id FROM users WHERE email=?", (req.email.lower(),)).fetchone()
    if existing:
        conn.close()
        raise HTTPException(400, "Email already registered")
    cur = conn.execute(
        "INSERT INTO users (email, name, password_hash, role, created_at) VALUES (?,?,?,?,?)",
        (req.email.lower(), req.name, hash_password(req.password), 'user', datetime.utcnow().isoformat())
    )
    uid = cur.lastrowid
    conn.commit()
    conn.close()
    log_audit(uid, "register", f"email={req.email}", request.client.host)
    token = create_jwt(uid, req.email.lower(), 'user')
    return {"token": token, "user": {"id": uid, "email": req.email.lower(), "name": req.name, "role": "user"}}


@app.post("/auth/login")
def login(req: LoginReq, request: Request):
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE email=? AND is_active=1", (req.email.lower(),)).fetchone()
    if not user or not verify_password(req.password, user["password_hash"]):
        conn.close()
        raise HTTPException(401, "Invalid credentials")
    conn.execute("UPDATE users SET last_login=? WHERE id=?", (datetime.utcnow().isoformat(), user["id"]))
    conn.commit()
    conn.close()
    log_audit(user["id"], "login", None, request.client.host)
    token = create_jwt(user["id"], user["email"], user["role"])
    return {"token": token, "user": {"id": user["id"], "email": user["email"], "name": user["name"], "role": user["role"]}}


@app.get("/auth/me")
def me(user=Depends(get_user)):
    conn = get_db()
    row = conn.execute("SELECT id, email, name, role, created_at, last_login FROM users WHERE id=?",
                       (user["user_id"],)).fetchone()
    conn.close()
    return dict(row) if row else {"error": "not found"}


@app.post("/auth/logout")
def logout(user=Depends(get_user)):
    log_audit(user["user_id"], "logout")
    return {"message": "logged out"}


# ════════════════════════════════════════════════════════════════════════════
# KEY MANAGEMENT
# ════════════════════════════════════════════════════════════════════════════

class CreateKeyReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=100)
    rate_limit_per_min: Optional[int] = None
    rate_limit_per_day: Optional[int] = None
    token_budget: Optional[int] = None
    expires_in_days: Optional[int] = None
    allowed_endpoints: Optional[str] = None


class PublicRegisterReq(BaseModel):
    email: str
    name: str


@app.post("/v1/keys/register", status_code=201)
def public_register_key(req: PublicRegisterReq, request: Request):
    """Public endpoint — Gemini-style key registration. Creates user if needed."""
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE email=?", (req.email.lower(),)).fetchone()
    if not user:
        # Auto-create user with random password
        random_pw = secrets.token_urlsafe(12)
        cur = conn.execute(
            "INSERT INTO users (email, name, password_hash, role, created_at) VALUES (?,?,?,?,?)",
            (req.email.lower(), req.name, hash_password(random_pw), 'user', datetime.utcnow().isoformat())
        )
        user_id = cur.lastrowid
    else:
        user_id = user["id"]

    key = generate_api_key()
    cur = conn.execute("""
        INSERT INTO api_keys (user_id, key, name, created_at, rate_limit_per_min, rate_limit_per_day, token_budget)
        VALUES (?,?,?,?,?,?,?)
    """, (user_id, key, f"{req.name}'s key", datetime.utcnow().isoformat(),
          settings.DEFAULT_RATE_PER_MIN, settings.DEFAULT_RATE_PER_DAY, settings.DEFAULT_TOKEN_BUDGET))
    kid = cur.lastrowid
    conn.commit()
    conn.close()
    log_audit(user_id, "public_key_registered", f"email={req.email}", request.client.host)
    return {
        "id": kid, "api_key": key, "name": req.name,
        "rate_limit_per_min": settings.DEFAULT_RATE_PER_MIN,
        "rate_limit_per_day": settings.DEFAULT_RATE_PER_DAY,
        "token_budget": settings.DEFAULT_TOKEN_BUDGET
    }


@app.post("/keys/create", status_code=201)
def create_key(req: CreateKeyReq, user=Depends(get_user)):
    key = generate_api_key()
    rpm = req.rate_limit_per_min or settings.DEFAULT_RATE_PER_MIN
    rpd = req.rate_limit_per_day or settings.DEFAULT_RATE_PER_DAY
    budget = req.token_budget or settings.DEFAULT_TOKEN_BUDGET
    expires = None
    if req.expires_in_days:
        expires = (datetime.utcnow() + timedelta(days=req.expires_in_days)).isoformat()
    endpoints = req.allowed_endpoints or "*"

    conn = get_db()
    cur = conn.execute("""
        INSERT INTO api_keys (user_id, key, name, created_at, rate_limit_per_min, rate_limit_per_day,
                              token_budget, expires_at, allowed_endpoints)
        VALUES (?,?,?,?,?,?,?,?,?)
    """, (user["user_id"], key, req.name, datetime.utcnow().isoformat(), rpm, rpd, budget, expires, endpoints))
    kid = cur.lastrowid
    conn.commit()
    conn.close()
    log_audit(user["user_id"], "key_created", f"id={kid}")
    return {"id": kid, "api_key": key, "name": req.name, "rate_limit_per_min": rpm,
            "rate_limit_per_day": rpd, "token_budget": budget, "expires_at": expires}


@app.get("/keys/list")
def list_keys(user=Depends(get_user)):
    conn = get_db()
    if user["role"] == "admin":
        rows = conn.execute("""
            SELECT ak.*, u.email as user_email, u.name as user_name
            FROM api_keys ak JOIN users u ON ak.user_id = u.id
            ORDER BY ak.created_at DESC
        """).fetchall()
    else:
        rows = conn.execute("SELECT * FROM api_keys WHERE user_id=? ORDER BY created_at DESC",
                            (user["user_id"],)).fetchall()
    conn.close()
    return {"keys": [dict(r) for r in rows]}


@app.delete("/keys/{key_id}")
def delete_key(key_id: int, user=Depends(get_user)):
    conn = get_db()
    if user["role"] == "admin":
        conn.execute("UPDATE api_keys SET is_active=0 WHERE id=?", (key_id,))
    else:
        conn.execute("UPDATE api_keys SET is_active=0 WHERE id=? AND user_id=?",
                     (key_id, user["user_id"]))
    conn.commit()
    conn.close()
    log_audit(user["user_id"], "key_revoked", f"id={key_id}")
    return {"message": "deactivated"}


class UpdateLimitsReq(BaseModel):
    rate_limit_per_min: Optional[int] = None
    rate_limit_per_day: Optional[int] = None
    token_budget: Optional[int] = None
    expires_in_days: Optional[int] = None


@app.put("/keys/{key_id}/limits")
def update_limits(key_id: int, req: UpdateLimitsReq, user=Depends(get_user)):
    conn = get_db()
    # Check ownership
    if user["role"] != "admin":
        row = conn.execute("SELECT user_id FROM api_keys WHERE id=?", (key_id,)).fetchone()
        if not row or row["user_id"] != user["user_id"]:
            conn.close()
            raise HTTPException(403, "Not your key")
    updates, params = [], []
    if req.rate_limit_per_min:
        updates.append("rate_limit_per_min=?")
        params.append(req.rate_limit_per_min)
    if req.rate_limit_per_day:
        updates.append("rate_limit_per_day=?")
        params.append(req.rate_limit_per_day)
    if req.token_budget:
        updates.append("token_budget=?")
        params.append(req.token_budget)
    if req.expires_in_days:
        updates.append("expires_at=?")
        params.append((datetime.utcnow() + timedelta(days=req.expires_in_days)).isoformat())
    if not updates:
        conn.close()
        raise HTTPException(400, "Nothing to update")
    params.append(key_id)
    conn.execute(f"UPDATE api_keys SET {', '.join(updates)} WHERE id=?", params)
    conn.commit()
    conn.close()
    log_audit(user["user_id"], "key_limits_updated", f"id={key_id}")
    return {"message": "updated"}


@app.get("/usage")
def get_usage(user=Depends(get_user), api_key: Optional[str] = None, limit: int = 100):
    conn = get_db()
    if user["role"] == "admin":
        if api_key:
            logs = conn.execute("SELECT * FROM usage_logs WHERE api_key=? ORDER BY timestamp DESC LIMIT ?",
                                (api_key, limit)).fetchall()
        else:
            logs = conn.execute("SELECT * FROM usage_logs ORDER BY timestamp DESC LIMIT ?", (limit,)).fetchall()
        summary = conn.execute("""
            SELECT api_key, COUNT(*) as total_requests,
                   SUM(prompt_tokens) as prompt_tokens, SUM(completion_tokens) as completion_tokens,
                   SUM(cost) as cost FROM usage_logs GROUP BY api_key
        """).fetchall()
    else:
        logs = conn.execute("SELECT * FROM usage_logs WHERE user_id=? ORDER BY timestamp DESC LIMIT ?",
                            (user["user_id"], limit)).fetchall()
        summary = conn.execute("""
            SELECT api_key, COUNT(*) as total_requests,
                   SUM(prompt_tokens) as prompt_tokens, SUM(completion_tokens) as completion_tokens,
                   SUM(cost) as cost FROM usage_logs WHERE user_id=? GROUP BY api_key
        """, (user["user_id"],)).fetchall()
    conn.close()
    return {"logs": [dict(r) for r in logs], "summary": [dict(r) for r in summary]}


@app.get("/usage/export")
def export_usage(user=Depends(get_user)):
    conn = get_db()
    if user["role"] == "admin":
        rows = conn.execute("SELECT * FROM usage_logs ORDER BY timestamp DESC").fetchall()
    else:
        rows = conn.execute("SELECT * FROM usage_logs WHERE user_id=? ORDER BY timestamp DESC",
                            (user["user_id"],)).fetchall()
    conn.close()
    if not rows:
        return Response("api_key,endpoint,prompt_tokens,completion_tokens,cost,timestamp\n",
                        media_type="text/csv")
    header = "id,api_key,user_id,endpoint,mode,prompt_tokens,completion_tokens,total_tokens,cost,latency_ms,status,timestamp\n"
    lines = [header]
    for r in rows:
        lines.append(f'{r["id"]},{r["api_key"]},{r["user_id"]},{r["endpoint"]},{r["mode"] or ""},'
                     f'{r["prompt_tokens"]},{r["completion_tokens"]},{r["total_tokens"]},'
                     f'{r["cost"]},{r["latency_ms"]},{r["status"]},{r["timestamp"]}\n')
    return Response("".join(lines), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=usage.csv"})


# ════════════════════════════════════════════════════════════════════════════
# EXTRACTION MODES
# ════════════════════════════════════════════════════════════════════════════

class ModeReq(BaseModel):
    name: str = Field(..., min_length=1, max_length=50)
    display_name: str
    description: Optional[str] = None
    system_prompt: str
    json_schema: Optional[str] = None
    temperature: float = 0.0
    max_tokens: int = 2048


@app.get("/v1/modes")
def list_modes_public():
    """Public: list active modes (for dropdown)."""
    conn = get_db()
    rows = conn.execute(
        "SELECT name, display_name, description FROM extraction_modes WHERE is_active=1"
    ).fetchall()
    conn.close()
    return {"modes": [dict(r) for r in rows]}


@app.get("/admin/modes")
def admin_list_modes(_=Depends(get_admin)):
    conn = get_db()
    rows = conn.execute("SELECT * FROM extraction_modes ORDER BY name").fetchall()
    conn.close()
    return {"modes": [dict(r) for r in rows]}


@app.get("/admin/modes/{name}")
def admin_get_mode(name: str, _=Depends(get_admin)):
    conn = get_db()
    row = conn.execute("SELECT * FROM extraction_modes WHERE name=?", (name,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "Mode not found")
    return dict(row)


@app.post("/admin/modes", status_code=201)
def admin_create_mode(req: ModeReq, admin=Depends(get_admin)):
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO extraction_modes (name, display_name, description, system_prompt,
                                          json_schema, temperature, max_tokens, created_at)
            VALUES (?,?,?,?,?,?,?,?)
        """, (req.name, req.display_name, req.description, req.system_prompt,
              req.json_schema, req.temperature, req.max_tokens, datetime.utcnow().isoformat()))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(400, "Mode name already exists")
    conn.close()
    log_audit(admin["user_id"], "mode_created", f"name={req.name}")
    return {"message": "created", "name": req.name}


@app.put("/admin/modes/{name}")
def admin_update_mode(name: str, req: ModeReq, admin=Depends(get_admin)):
    conn = get_db()
    conn.execute("""
        UPDATE extraction_modes SET display_name=?, description=?, system_prompt=?,
                                     json_schema=?, temperature=?, max_tokens=?, updated_at=?
        WHERE name=?
    """, (req.display_name, req.description, req.system_prompt, req.json_schema,
          req.temperature, req.max_tokens, datetime.utcnow().isoformat(), name))
    conn.commit()
    conn.close()
    log_audit(admin["user_id"], "mode_updated", f"name={name}")
    return {"message": "updated"}


@app.delete("/admin/modes/{name}")
def admin_delete_mode(name: str, hard: bool = False, admin=Depends(get_admin)):
    conn = get_db()
    if hard:
        conn.execute("DELETE FROM extraction_modes WHERE name=?", (name,))
    else:
        conn.execute("UPDATE extraction_modes SET is_active=0 WHERE name=?", (name,))
    conn.commit()
    conn.close()
    log_audit(admin["user_id"], "mode_deleted", f"name={name} hard={hard}")
    return {"message": "deleted"}


def get_mode_config(mode_name: str) -> Optional[dict]:
    conn = get_db()
    row = conn.execute("SELECT * FROM extraction_modes WHERE name=? AND is_active=1",
                       (mode_name,)).fetchone()
    conn.close()
    return dict(row) if row else None


# ════════════════════════════════════════════════════════════════════════════
# AI ROUTES — CHAT COMPLETIONS (OpenAI compatible)
# ════════════════════════════════════════════════════════════════════════════

class ChatMessage(BaseModel):
    role: str
    content: Any

class ChatReq(BaseModel):
    messages: List[ChatMessage]
    max_tokens: Optional[int] = Field(2048, ge=1, le=16384)
    temperature: Optional[float] = Field(0.0, ge=0, le=2)
    top_p: Optional[float] = Field(1.0, ge=0, le=1)
    enable_thinking: Optional[bool] = False
    stream: Optional[bool] = False


@app.post("/v1/chat/completions")
async def chat_completions(req: ChatReq, request: Request, info=Depends(get_key_info)):
    if len(req.messages) > settings.MAX_MESSAGES:
        raise HTTPException(400, f"Too many messages (max {settings.MAX_MESSAGES})")

    if req.stream:
        async def gen():
            tf = ThinkFilter()
            pt = ct = 0
            start = time.time()
            try:
                async for chunk in vllm_stream([m.dict() for m in req.messages],
                                               req.max_tokens, req.temperature, req.enable_thinking):
                    clean = tf.feed(chunk)
                    if clean:
                        ct += len(clean.split())
                        yield f"data: {json.dumps({'choices':[{'delta':{'content':clean}}]})}\n\n"
                tail = tf.flush()
                if tail:
                    yield f"data: {json.dumps({'choices':[{'delta':{'content':tail}}]})}\n\n"
                yield "data: [DONE]\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'error': str(e)})}\n\n"
            finally:
                latency = int((time.time() - start) * 1000)
                log_usage(info["key"], info["user_id"], "/v1/chat/completions",
                          pt, ct, latency, "success", None, request.client.host)
        return StreamingResponse(gen(), media_type="text/event-stream")

    start = time.time()
    try:
        result = await vllm_chat([m.dict() for m in req.messages],
                                 req.max_tokens, req.temperature, req.enable_thinking)
        latency = int((time.time() - start) * 1000)
        u = result.get("usage", {})
        # Strip <think> from final response
        for choice in result.get("choices", []):
            if choice.get("message", {}).get("content"):
                choice["message"]["content"] = strip_think(choice["message"]["content"])
        log_usage(info["key"], info["user_id"], "/v1/chat/completions",
                  u.get("prompt_tokens", 0), u.get("completion_tokens", 0),
                  latency, "success", None, request.client.host)
        return result
    except httpx.HTTPStatusError as e:
        log_usage(info["key"], info["user_id"], "/v1/chat/completions", 0, 0,
                  int((time.time() - start) * 1000), "error", str(e)[:200], request.client.host)
        raise HTTPException(502, f"vLLM error: {e.response.text[:200]}")
    except Exception as e:
        log_usage(info["key"], info["user_id"], "/v1/chat/completions", 0, 0,
                  int((time.time() - start) * 1000), "error", str(e)[:200], request.client.host)
        raise HTTPException(500, str(e)[:200])


# ════════════════════════════════════════════════════════════════════════════
# /v1/generate — simple text generation
# ════════════════════════════════════════════════════════════════════════════

class GenerateReq(BaseModel):
    prompt: str
    system: Optional[str] = None
    mode: Optional[str] = None
    max_tokens: Optional[int] = 2048
    temperature: Optional[float] = 0.7
    enable_thinking: Optional[bool] = False


@app.post("/v1/generate")
async def generate(req: GenerateReq, request: Request, info=Depends(get_key_info)):
    start = time.time()
    try:
        result = await vllm_chat([{"role": "user", "content": req.prompt}],
                                 req.max_tokens, req.temperature, req.enable_thinking)
        latency = int((time.time() - start) * 1000)
        u = result.get("usage", {})
        content = strip_think(result["choices"][0]["message"]["content"])
        log_usage(info["key"], info["user_id"], "/v1/generate",
                  u.get("prompt_tokens", 0), u.get("completion_tokens", 0),
                  latency, "success", None, request.client.host)
        return {"text": content, "usage": u, "latency_ms": latency}
    except Exception as e:
        log_usage(info["key"], info["user_id"], "/v1/generate", 0, 0,
                  int((time.time() - start) * 1000), "error", str(e)[:200], request.client.host)
        raise HTTPException(500, str(e)[:200])


@app.post("/v1/generate/stream")
async def generate_stream(req: GenerateReq, request: Request, info=Depends(get_key_info)):
    # Build messages with optional system prompt and mode
    messages = []
    mode_config = get_mode_config(req.mode) if req.mode and req.mode != "free" else None
    if mode_config and mode_config.get("system_prompt"):
        messages.append({"role": "system", "content": mode_config["system_prompt"]})
    elif req.system:
        messages.append({"role": "system", "content": req.system})
    messages.append({"role": "user", "content": req.prompt})

    max_tok = mode_config["max_tokens"] if mode_config else (req.max_tokens or 2048)
    temp    = mode_config["temperature"] if mode_config else (req.temperature or 0.7)

    _SSE_HEADERS = {
        "Cache-Control": "no-cache, no-transform",
        "Connection": "keep-alive",
        "X-Accel-Buffering": "no",
    }

    async def gen():
        start = time.time()
        pt = ct = 0
        try:
            async for event_type, text, usage in vllm_stream_events(
                messages, max_tok, temp, req.enable_thinking
            ):
                if event_type == "delta":
                    yield f"data: {json.dumps({'event': 'delta', 'text': text})}\n\n"
                elif event_type == "thinking":
                    yield f"data: {json.dumps({'event': 'thinking', 'text': text})}\n\n"
                elif event_type == "done":
                    pt = (usage or {}).get("prompt_tokens", 0)
                    ct = (usage or {}).get("completion_tokens", 0)
                    yield f"data: {json.dumps({'event': 'done', 'finish_reason': text, 'usage': {'input_tokens': pt, 'output_tokens': ct, 'total_tokens': pt + ct}, 'model': settings.MODEL_DISPLAY_NAME})}\n\n"
                elif event_type == "error":
                    yield f"data: {json.dumps({'event': 'error', 'error': text})}\n\n"
        finally:
            latency = int((time.time() - start) * 1000)
            log_usage(info["key"], info["user_id"], "/v1/generate/stream",
                      pt, ct, latency, "success", None, request.client.host)

    return StreamingResponse(gen(), media_type="text/event-stream", headers=_SSE_HEADERS)


# ════════════════════════════════════════════════════════════════════════════
# /v1/ask — Universal endpoint with intent detection
# ════════════════════════════════════════════════════════════════════════════

@app.post("/v1/ask")
async def ask(
    request: Request,
    prompt: str = Form(...),
    file: Optional[UploadFile] = File(None),
    mode: Optional[str] = Form(None),
    enable_thinking: bool = Form(False),
    info=Depends(get_key_info)
):
    """Universal endpoint: text + optional file + auto intent detection."""
    start = time.time()
    file_data, file_mime, file_name = None, None, ""
    if file:
        file_data = await file.read()
        if len(file_data) > settings.MAX_UPLOAD_MB * 1024 * 1024:
            raise HTTPException(400, f"File too large (max {settings.MAX_UPLOAD_MB}MB)")
        file_mime = file.content_type
        file_name = file.filename or ""

    # Auto-detect intent if file present and mode not specified
    detected_mode = mode or detect_intent(prompt, has_file=bool(file_data))
    mode_config = get_mode_config(detected_mode) if detected_mode else None

    # Build messages
    if file_data:
        ext = (file_name.split(".")[-1] if "." in file_name else "").lower()
        is_image_like = (file_mime and file_mime.startswith("image/")) or ext in ("png", "jpg", "jpeg", "webp", "pdf")
        if is_image_like:
            images = await file_to_images_b64(file_data, file_mime, file_name)
            content = [{"type": "image_url", "image_url": img} for img in images]
            sys_prompt = mode_config["system_prompt"] if mode_config else "Analyze the image and respond to the user prompt."
            content.append({"type": "text", "text": prompt})
            messages = [{"role": "system", "content": sys_prompt}, {"role": "user", "content": content}]
            max_tokens = mode_config["max_tokens"] if mode_config else 4096
            temperature = mode_config["temperature"] if mode_config else 0.0
        else:
            # Text-based file
            text_content = file_to_text(file_data, file_name)
            user_prompt = f"{prompt}\n\n--- File Content ({file_name}) ---\n{text_content[:50000]}"
            messages = [{"role": "user", "content": user_prompt}]
            max_tokens = mode_config["max_tokens"] if mode_config else 4096
            temperature = mode_config["temperature"] if mode_config else 0.0
    else:
        messages = [{"role": "user", "content": prompt}]
        max_tokens = 2048
        temperature = 0.0

    try:
        result = await vllm_chat(messages, max_tokens, temperature, enable_thinking)
        latency = int((time.time() - start) * 1000)
        u = result.get("usage", {})
        content = strip_think(result["choices"][0]["message"]["content"])

        # Try parse JSON if mode used
        parsed = None
        if mode_config:
            try:
                # Strip markdown code fences
                cleaned = re.sub(r"^```(?:json)?\s*|\s*```$", "", content.strip(), flags=re.MULTILINE)
                parsed = json.loads(cleaned)
            except Exception:
                parsed = None

        log_usage(info["key"], info["user_id"], "/v1/ask",
                  u.get("prompt_tokens", 0), u.get("completion_tokens", 0),
                  latency, "success", None, request.client.host, detected_mode)
        return {
            "result": content,
            "parsed": parsed,
            "mode": detected_mode,
            "usage": u,
            "latency_ms": latency
        }
    except Exception as e:
        log_usage(info["key"], info["user_id"], "/v1/ask", 0, 0,
                  int((time.time() - start) * 1000), "error", str(e)[:200],
                  request.client.host, detected_mode)
        raise HTTPException(500, str(e)[:200])


# ════════════════════════════════════════════════════════════════════════════
# /v1/vision/analyze
# ════════════════════════════════════════════════════════════════════════════

@app.post("/v1/vision/analyze")
async def vision_analyze(
    request: Request,
    file: UploadFile = File(...),
    prompt: str = Form(default="Extract all data as structured JSON."),
    mode: Optional[str] = Form(None),
    enable_thinking: bool = Form(False),
    max_tokens: int = Form(4096),
    info=Depends(get_key_info)
):
    start = time.time()
    data = await file.read()
    if len(data) > settings.MAX_UPLOAD_MB * 1024 * 1024:
        raise HTTPException(400, f"File too large (max {settings.MAX_UPLOAD_MB}MB)")

    mode_config = get_mode_config(mode) if mode else None
    images = await file_to_images_b64(data, file.content_type or "image/jpeg", file.filename or "")

    sys_prompt = mode_config["system_prompt"] if mode_config else "Analyze the image."
    user_content = [{"type": "image_url", "image_url": img} for img in images]
    user_content.append({"type": "text", "text": prompt})
    messages = [{"role": "system", "content": sys_prompt}, {"role": "user", "content": user_content}]

    try:
        result = await vllm_chat(messages, max_tokens, mode_config["temperature"] if mode_config else 0.0, enable_thinking)
        latency = int((time.time() - start) * 1000)
        u = result.get("usage", {})
        content = strip_think(result["choices"][0]["message"]["content"])

        parsed = None
        if mode_config:
            try:
                cleaned = re.sub(r"^```(?:json)?\s*|\s*```$", "", content.strip(), flags=re.MULTILINE)
                parsed = json.loads(cleaned)
            except Exception:
                pass

        log_usage(info["key"], info["user_id"], "/v1/vision/analyze",
                  u.get("prompt_tokens", 0), u.get("completion_tokens", 0),
                  latency, "success", None, request.client.host, mode)
        return {"result": content, "parsed": parsed, "mode": mode, "usage": u,
                "pages_processed": len(images), "latency_ms": latency}
    except Exception as e:
        log_usage(info["key"], info["user_id"], "/v1/vision/analyze", 0, 0,
                  int((time.time() - start) * 1000), "error", str(e)[:200],
                  request.client.host, mode)
        raise HTTPException(500, str(e)[:200])


# ════════════════════════════════════════════════════════════════════════════
# /v1/vision/analyze/stream — SSE streaming vision analysis
# ════════════════════════════════════════════════════════════════════════════

@app.post("/v1/vision/analyze/stream")
async def vision_analyze_stream(
    request: Request,
    file: UploadFile = File(...),
    prompt: str = Form(default="Describe this image/document in detail."),
    mode: Optional[str] = Form(None),
    enable_thinking: bool = Form(False),
    max_tokens: int = Form(4096),
    temperature: float = Form(0.0),
    info=Depends(get_key_info),
):
    data = await file.read()
    if len(data) > settings.MAX_UPLOAD_MB * 1024 * 1024:
        raise HTTPException(400, f"File too large (max {settings.MAX_UPLOAD_MB}MB)")

    try:
        images = await file_to_images_b64(data, file.content_type or "image/jpeg", file.filename or "")
    except HTTPException as e:
        async def err_gen():
            yield f"data: {json.dumps({'event': 'error', 'error': e.detail})}\n\n"
        return StreamingResponse(err_gen(), media_type="text/event-stream")

    mode_config = get_mode_config(mode) if mode and mode != "free" else None
    sys_prompt = (mode_config["system_prompt"] if mode_config
                  else "Analyze the image/document and respond to the user's request.")
    user_content = [{"type": "image_url", "image_url": img} for img in images]
    user_content.append({"type": "text", "text": prompt})
    messages = [{"role": "system", "content": sys_prompt}, {"role": "user", "content": user_content}]
    temp    = mode_config["temperature"] if mode_config else temperature
    max_tok = mode_config["max_tokens"]  if mode_config else max_tokens

    _SSE_HEADERS = {
        "Cache-Control": "no-cache, no-transform",
        "Connection": "keep-alive",
        "X-Accel-Buffering": "no",
    }

    async def gen():
        start = time.time()
        pt = ct = 0
        yield f"data: {json.dumps({'event': 'meta', 'pages': len(images), 'mode': mode or 'free'})}\n\n"
        try:
            async for event_type, text, usage in vllm_stream_events(
                messages, max_tok, temp, enable_thinking
            ):
                if event_type == "delta":
                    yield f"data: {json.dumps({'event': 'delta', 'text': text})}\n\n"
                elif event_type == "thinking":
                    yield f"data: {json.dumps({'event': 'thinking', 'text': text})}\n\n"
                elif event_type == "done":
                    pt = (usage or {}).get("prompt_tokens", 0)
                    ct = (usage or {}).get("completion_tokens", 0)
                    yield f"data: {json.dumps({'event': 'done', 'finish_reason': text, 'pages': len(images), 'mode': mode or 'free', 'usage': {'input_tokens': pt, 'output_tokens': ct, 'total_tokens': pt + ct}, 'model': settings.MODEL_DISPLAY_NAME})}\n\n"
                elif event_type == "error":
                    yield f"data: {json.dumps({'event': 'error', 'error': text})}\n\n"
        finally:
            latency = int((time.time() - start) * 1000)
            log_usage(info["key"], info["user_id"], "/v1/vision/analyze/stream",
                      pt, ct, latency, "success", None, request.client.host, mode)

    return StreamingResponse(gen(), media_type="text/event-stream", headers=_SSE_HEADERS)


# ════════════════════════════════════════════════════════════════════════════
# AUDIO ENDPOINTS
# ════════════════════════════════════════════════════════════════════════════

@app.post("/v1/audio/transcriptions")
async def transcribe(
    request: Request,
    file: UploadFile = File(...),
    language: Optional[str] = Form(None),
    info=Depends(get_key_info)
):
    data = await file.read()
    if len(data) > settings.MAX_UPLOAD_MB * 1024 * 1024:
        raise HTTPException(400, "File too large")
    start = time.time()
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".audio") as tmp:
            tmp.write(data)
            tmp_path = tmp.name
        model = get_whisper()
        segments, info_obj = model.transcribe(tmp_path, language=language, beam_size=5)
        text = " ".join([s.text for s in segments]).strip()
        latency = int((time.time() - start) * 1000)
        log_usage(info["key"], info["user_id"], "/v1/audio/transcriptions",
                  0, len(text.split()), latency, "success", None, request.client.host)
        return {"text": text, "language": info_obj.language, "duration": info_obj.duration,
                "latency_ms": latency}
    except Exception as e:
        log_usage(info["key"], info["user_id"], "/v1/audio/transcriptions", 0, 0,
                  int((time.time() - start) * 1000), "error", str(e)[:200], request.client.host)
        raise HTTPException(500, str(e)[:200])
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


@app.post("/v1/audio/chat")
async def audio_chat(
    request: Request,
    file: UploadFile = File(...),
    language: Optional[str] = Form(None),
    enable_thinking: bool = Form(False),
    info=Depends(get_key_info)
):
    data = await file.read()
    if len(data) > settings.MAX_UPLOAD_MB * 1024 * 1024:
        raise HTTPException(400, "File too large")
    start = time.time()
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".audio") as tmp:
            tmp.write(data)
            tmp_path = tmp.name
        model = get_whisper()
        segments, info_obj = model.transcribe(tmp_path, language=language, beam_size=5)
        user_text = " ".join([s.text for s in segments]).strip()
        result = await vllm_chat([{"role": "user", "content": user_text}], 2048, 0.0, enable_thinking)
        latency = int((time.time() - start) * 1000)
        u = result.get("usage", {})
        response_text = strip_think(result["choices"][0]["message"]["content"])
        log_usage(info["key"], info["user_id"], "/v1/audio/chat",
                  u.get("prompt_tokens", 0), u.get("completion_tokens", 0),
                  latency, "success", None, request.client.host)
        return {"transcription": user_text, "language": info_obj.language,
                "response": response_text, "usage": u, "latency_ms": latency}
    except Exception as e:
        log_usage(info["key"], info["user_id"], "/v1/audio/chat", 0, 0,
                  int((time.time() - start) * 1000), "error", str(e)[:200], request.client.host)
        raise HTTPException(500, str(e)[:200])
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


# ════════════════════════════════════════════════════════════════════════════
# /v1/quick — Browser GET test
# ════════════════════════════════════════════════════════════════════════════

@app.get("/v1/quick")
async def quick(q: str = Query(...), info=Depends(get_key_info)):
    start = time.time()
    try:
        result = await vllm_chat([{"role": "user", "content": q}], 1024, 0.0)
        u = result.get("usage", {})
        content = strip_think(result["choices"][0]["message"]["content"])
        log_usage(info["key"], info["user_id"], "/v1/quick",
                  u.get("prompt_tokens", 0), u.get("completion_tokens", 0),
                  int((time.time() - start) * 1000), "success")
        return {"answer": content}
    except Exception as e:
        log_usage(info["key"], info["user_id"], "/v1/quick", 0, 0, 0, "error", str(e)[:200])
        raise HTTPException(500, str(e)[:200])


# ════════════════════════════════════════════════════════════════════════════
# ADMIN ROUTES
# ════════════════════════════════════════════════════════════════════════════

@app.get("/admin/users")
def admin_users(_=Depends(get_admin)):
    conn = get_db()
    rows = conn.execute("""
        SELECT u.id, u.email, u.name, u.role, u.is_active, u.created_at, u.last_login,
               COUNT(DISTINCT ak.id) as key_count,
               COUNT(ul.id) as request_count,
               COALESCE(SUM(ul.cost), 0) as total_cost
        FROM users u
        LEFT JOIN api_keys ak ON ak.user_id = u.id
        LEFT JOIN usage_logs ul ON ul.user_id = u.id
        GROUP BY u.id ORDER BY u.created_at DESC
    """).fetchall()
    conn.close()
    return {"users": [dict(r) for r in rows]}


@app.post("/admin/users/{user_id}/toggle")
def admin_toggle_user(user_id: int, admin=Depends(get_admin)):
    if user_id == admin["user_id"]:
        raise HTTPException(400, "Cannot toggle yourself")
    conn = get_db()
    conn.execute("UPDATE users SET is_active = 1 - is_active WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    log_audit(admin["user_id"], "user_toggle", f"target={user_id}")
    return {"message": "toggled"}


@app.delete("/admin/users/{user_id}")
def admin_delete_user(user_id: int, admin=Depends(get_admin)):
    if user_id == admin["user_id"]:
        raise HTTPException(400, "Cannot delete yourself")
    conn = get_db()
    conn.execute("UPDATE users SET is_active=0 WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    log_audit(admin["user_id"], "user_deleted", f"target={user_id}")
    return {"message": "deleted"}


@app.get("/admin/stats")
def admin_stats(_=Depends(get_admin)):
    conn = get_db()
    yesterday = (datetime.utcnow() - timedelta(days=1)).isoformat()
    stats = {
        "users": conn.execute("SELECT COUNT(*) c FROM users WHERE is_active=1").fetchone()["c"],
        "keys": conn.execute("SELECT COUNT(*) c FROM api_keys WHERE is_active=1").fetchone()["c"],
        "total_requests": conn.execute("SELECT COUNT(*) c FROM usage_logs").fetchone()["c"],
        "total_cost": conn.execute("SELECT COALESCE(SUM(cost),0) c FROM usage_logs").fetchone()["c"],
        "total_tokens": conn.execute("SELECT COALESCE(SUM(total_tokens),0) c FROM usage_logs").fetchone()["c"],
        "requests_24h": conn.execute("SELECT COUNT(*) c FROM usage_logs WHERE timestamp>?", (yesterday,)).fetchone()["c"],
        "errors_24h": conn.execute("SELECT COUNT(*) c FROM usage_logs WHERE timestamp>? AND status='error'",
                                    (yesterday,)).fetchone()["c"],
    }
    by_endpoint = conn.execute("""
        SELECT endpoint, COUNT(*) as count, SUM(cost) as cost, AVG(latency_ms) as avg_latency
        FROM usage_logs GROUP BY endpoint
    """).fetchall()
    by_mode = conn.execute("""
        SELECT mode, COUNT(*) as count FROM usage_logs WHERE mode IS NOT NULL GROUP BY mode
    """).fetchall()
    conn.close()
    stats["by_endpoint"] = [dict(r) for r in by_endpoint]
    stats["by_mode"] = [dict(r) for r in by_mode]
    return stats


@app.get("/admin/audit")
def admin_audit(_=Depends(get_admin), limit: int = 100):
    conn = get_db()
    rows = conn.execute("SELECT * FROM audit_logs ORDER BY timestamp DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return {"audit_logs": [dict(r) for r in rows]}


@app.get("/admin/logs")
def admin_logs(_=Depends(get_admin), file: str = "backend", lines: int = 200):
    """Read last N lines from a log file. file: backend | requests | ai | errors"""
    allowed = {"backend": "backend.log", "requests": "requests.log",
               "ai": "ai.log", "errors": "errors.log"}
    fname = allowed.get(file)
    if not fname:
        raise HTTPException(400, f"Unknown log file. Choose from: {list(allowed.keys())}")
    path = _LOG_DIR / fname
    if not path.exists():
        return {"file": file, "lines": []}
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        all_lines = f.readlines()
    tail = [l.rstrip() for l in all_lines[-lines:]]
    return {"file": file, "path": str(path), "total_lines": len(all_lines), "lines": tail}


# ════════════════════════════════════════════════════════════════════════════
# FRONTEND SPA FALLBACK
# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
# EXTRA AUTH ROUTES — must be registered BEFORE the spa_fallback catch-all
# ════════════════════════════════════════════════════════════════════════════

@app.get("/auth/my-stats")
def my_stats_route(user=Depends(get_user)):
    conn = get_db()
    uid = user["user_id"]
    total_keys   = conn.execute("SELECT COUNT(*) FROM api_keys WHERE user_id=?", (uid,)).fetchone()[0]
    total_req    = conn.execute("SELECT COALESCE(SUM(total_requests),0) FROM api_keys WHERE user_id=?", (uid,)).fetchone()[0]
    total_tokens = conn.execute("SELECT COALESCE(SUM(tokens_used),0) FROM api_keys WHERE user_id=?", (uid,)).fetchone()[0]
    total_cost   = conn.execute("SELECT COALESCE(SUM(total_cost),0) FROM api_keys WHERE user_id=?", (uid,)).fetchone()[0]
    conn.close()
    return {"total_keys": total_keys, "total_requests": total_req, "total_tokens": total_tokens, "total_cost": total_cost}


@app.get("/auth/stats")
def admin_stats_route(_=Depends(get_admin)):
    conn = get_db()
    total_users  = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    total_keys   = conn.execute("SELECT COUNT(*) FROM api_keys WHERE is_active=1").fetchone()[0]
    total_req    = conn.execute("SELECT COALESCE(SUM(total_requests),0) FROM api_keys").fetchone()[0]
    total_tokens = conn.execute("SELECT COALESCE(SUM(tokens_used),0) FROM api_keys").fetchone()[0]
    total_cost   = conn.execute("SELECT COALESCE(SUM(total_cost),0) FROM api_keys").fetchone()[0]
    conn.close()
    return {"total_users": total_users, "total_keys": total_keys, "total_requests": total_req,
            "total_tokens": total_tokens, "total_cost": total_cost}


@app.get("/auth/users")
def list_users_route(_=Depends(get_admin)):
    conn = get_db()
    rows = conn.execute(
        "SELECT id, email, name, role, is_active, created_at, last_login FROM users ORDER BY created_at DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.delete("/auth/users/{user_id}")
def delete_user_route(user_id: int, _=Depends(get_admin)):
    conn = get_db()
    conn.execute("UPDATE users SET is_active=0 WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    return {"message": "deactivated"}


# ════════════════════════════════════════════════════════════════════════════
# FRONTEND SPA FALLBACK — must be LAST (catch-all)
# ════════════════════════════════════════════════════════════════════════════

if os.path.isdir(os.path.join(settings.DIST_DIR, "assets")):
    app.mount("/assets", StaticFiles(directory=os.path.join(settings.DIST_DIR, "assets")), name="assets")


@app.get("/{full_path:path}", include_in_schema=False)
async def spa_fallback(full_path: str):
    api_prefixes = ("v1/", "auth/", "keys", "admin/", "usage", "health", "docs", "openapi", "redoc")
    if any(full_path.startswith(p) for p in api_prefixes):
        raise HTTPException(404, "Not found")
    index = os.path.join(settings.DIST_DIR, "index.html")
    if os.path.isfile(index):
        return FileResponse(index)
    return HTMLResponse(
        f"<h1>{settings.PROJECT_NAME}</h1>"
        f"<p>Frontend not built. API at <a href='/docs'>/docs</a></p>"
    )

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=1111, reload=False, log_level="info")
